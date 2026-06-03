#!/usr/bin/env python3
"""
Recalculate all formulas in an Excel file using LibreOffice.
"""

import sys
import os
import json
import subprocess
import tempfile
from pathlib import Path

def recalculate_formulas(excel_file, timeout=30):
    """Recalculate formulas in Excel file using LibreOffice."""
    excel_path = Path(excel_file).resolve()
    
    if not excel_path.exists():
        return {
            "status": "error",
            "message": f"File not found: {excel_path}"
        }
    
    # Use LibreOffice to open and save the file (forces recalculation)
    output_path = excel_path
    
    try:
        # Open with LibreOffice in headless mode and save (triggers formula recalc)
        cmd = [
            'soffice',
            '--headless',
            '--convert-to', 'xlsx',
            '--outdir', str(excel_path.parent),
            str(excel_path)
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        
        if result.returncode == 0:
            # Read the file to check for errors
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=False)
            
            errors_found = {}
            error_count = 0
            formula_count = 0
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_count += 1
                            # Check if cell is showing an error
                            if cell.data_type == 'e':  # Error type
                                error_code = cell.value
                                if error_code not in errors_found:
                                    errors_found[error_code] = {
                                        "count": 0,
                                        "locations": []
                                    }
                                errors_found[error_code]["count"] += 1
                                errors_found[error_code]["locations"].append(f"{sheet}!{cell.coordinate}")
                                error_count += 1
            
            wb.close()
            
            if error_count > 0:
                return {
                    "status": "errors_found",
                    "total_errors": error_count,
                    "total_formulas": formula_count,
                    "error_summary": errors_found
                }
            else:
                return {
                    "status": "success",
                    "total_errors": 0,
                    "total_formulas": formula_count
                }
        else:
            return {
                "status": "error",
                "message": f"LibreOffice conversion failed: {result.stderr}",
                "stderr": result.stderr
            }
            
    except subprocess.TimeoutExpired:
        return {
            "status": "error",
            "message": f"Recalculation timed out after {timeout}s"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    result = recalculate_formulas(excel_file, timeout)
    print(json.dumps(result, indent=2))
    
    if result.get("status") == "error":
        sys.exit(1)
    elif result.get("status") == "errors_found":
        sys.exit(2)
